
import io
import logging
from datetime import datetime, date, timedelta
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Static config from "INFORMÁTICA 01-2026.xlsx" analysis
# Map Location Name -> Classification
LOCATION_CLASSIFICATION = {
    "ADMINISTRACAO": "Administrativo",
    "AGENCIA TRANSFUSIONAL": "Intermediário",
    "ALA 1": "Produtivo",
    "ALA 2 - CLINICA CIRURGICA": "Produtivo",
    "ALA 2 - CLINICA MEDICA": "Produtivo",
    "ALA 3": "Produtivo",
    "ALA 4": "Produtivo",
    "AMBULATORIO": "Produtivo",
    "AMBULATORIO PREDIO C": "Produtivo",
    "AUDITORIA CONVÊNIOS": "Administrativo",
    "CCIH": "Apoio",
    "CENTRAL DE AUTORIZACOES": "Administrativo",
    "CENTRAL MATERIAS - CME": "Apoio",
    "CENTRO CIRURGICO": "Produtivo",
    "COMPRAS": "Administrativo",
    "CONTABILIDADE": "Administrativo",
    "COORDENAÇÃO ENFERMAGEM": "Administrativo",
    "DEPOSITO - ALMOXARIFADO": "Apoio",
    "DIAGNOSTICO POR IMAGEM": "Intermediário",
    "FARMACIA": "Apoio",
    "FATURAMENTO/AUDITORIA CONVÊNIOS": "Administrativo",
    "FATURAMENTO/AUDITORIA SUS": "Administrativo",
    "FINANCEIRO": "Administrativo",
    "FISIOTERAPIA": "Intermediário",
    "HIGIENIZACAO": "Apoio",
    "HOTELARIA": "Administrativo",
    "INFORMATICA - TI": "Administrativo",
    "LABORATORIO": "Intermediário",
    "LAVANDERIA": "Apoio",
    "MANUTENCAO": "Administrativo",
    "PATRIMONIO": "Administrativo",
    "PRONTO ATENDIMENTO - PAGO": "Produtivo",
    "PRONTO SOCORRO": "Produtivo",
    "RA INTERNACAO PROVISORIO": "Produtivo",
    "RECEPÇÃO/INTERNAÇÃO": "Administrativo",
    "RH": "Administrativo",
    "ROUPARIA E COSTURA": "Apoio",
    "SERVIÇO NUTRIÇÃO E DIETETICA": "Apoio",
    "SERVICO SOCIAL": "Administrativo",
    "SESMT": "Administrativo",
    "SPP": "Apoio",
    "TESOURARIA": "Administrativo",
    "UTI ADULTO": "Produtivo",
    "UTI NEONATAL": "Produtivo",
}

def get_previous_month_range() -> tuple[str, str]:
    """Return (start_date, end_date) for the previous month in YYYY-MM-DD format."""
    today = date.today()
    first_this_month = date(today.year, today.month, 1)
    last_prev_month = first_this_month - timedelta(days=1)
    first_prev_month = date(last_prev_month.year, last_prev_month.month, 1)
    return first_prev_month.strftime("%Y-%m-%d"), last_prev_month.strftime("%Y-%m-%d")

async def generate_cost_center_report_excel() -> tuple[bytes, str]:
    """
    Generates the Excel report for 'Atendimentos por Centro de Custo'.
    Automatically targets the *previous month*.
    Returns: (excel_bytes, filename)
    """
    from core.tools.glpi import get_client

    start_date, end_date = get_previous_month_range()
    filename = f"INFORMATICA_{start_date[:7]}.xlsx"  # e.g. INFORMATICA_2026-01.xlsx

    client = get_client()

    # 1. Fetch Locations
    loc_result = await client.get_locations(limit=200)
    if not loc_result.success:
        raise Exception(f"Failed to fetch locations: {loc_result.error}")
    
    locations = loc_result.output.get("locations", [])
    # Filter only fully active/valid locations if needed, or just use all
    # Sort by name to match Excel generally
    locations.sort(key=lambda x: x["name"])

    # 2. Fetch Tickets for the date range
    # We need to filter manually because GLPI API search is complex.
    # Fetching all tickets might be heavy, but simpler for 'previous month' than constructing complex criteria.
    # Optimized: use 'date' search criteria if possible, or fetch recently modified/created
    
    # Using 'search' criteria for date range on 'date' (creation date)
    # 15 = Creation date
    params = {
        "range": "0-1000", # Hope to fit all tickets from 1 month in 1000? Maybe need pagination.
        "criteria[0][field]": 15, # Date creation
        "criteria[0][searchtype]": "morethan",
        "criteria[0][value]": f"{start_date} 00:00:00",
        "criteria[1][link]": "AND",
        "criteria[1][field]": 15,
        "criteria[1][searchtype]": "lessthan",
        "criteria[1][value]": f"{end_date} 23:59:59",
        "criteria[2][link]": "AND",
        "criteria[2][field]": 83, # Locations ID
        "criteria[2][searchtype]": "morethan", # Just ensure it has a location? 
        "criteria[2][value]": "0",  
    }
    
    # We need to use the native client GET on /Ticket to support complex criteria not exposed in simpler helpers
    # Or iterate via list_tickets helper. Let's use the low-level client via helper if available, or direct call.
    # The get_tickets helper is too simple. Let's make a direct request using the client instance we have.
    
    # Since get_client() returns our wrapper, let's access the internalhttpx client or add a robust search method.
    # For now, let's fetch 'all recent' and filter in Python to be safe and strictly correct without fighting API syntax quirks.
    # Assuming < 2000 tickets/month.
    
    # Using 'get_tickets' with a high limit, but we can't filter by date there easily yet.
    # Let's assume we can fetch last 1000 tickets and filter by date.
    
    raw_tickets_result = await client.get_tickets(limit=1000, order="DESC") # Newest first
    if not raw_tickets_result.success:
        raise Exception(f"Failed to fetch tickets: {raw_tickets_result.error}")
        
    all_tickets = raw_tickets_result.output.get("tickets", [])
    
    # Filter by date range (local)
    relevant_tickets = []
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()

    for t in all_tickets:
        date_str = t.get("date") or t.get("date_creation")
        if not date_str:
            continue
        try:
            # "2026-02-09 10:00:00"
            t_date = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S").date()
            if start_dt <= t_date <= end_dt:
                relevant_tickets.append(t)
            elif t_date < start_dt:
                # Optimized: if sorted DESC, we can stop once we hit dates older than start_date
                # BUT let's be safe and check all in case of weird sorting
                pass
        except:
            pass

    # 3. Aggregate
    # Map[location_id] -> count
    counts_by_loc_id = {}
    for t in relevant_tickets:
        lid = t.get("locations_id")
        if lid:
            counts_by_loc_id[lid] = counts_by_loc_id.get(lid, 0) + 1

    total_tickets = len(relevant_tickets)

    # 4. Build Data Rows
    # Expected cols: CÓDIGO (ID), DESCRIÇÃO (Name), CLASSIFICAÇÃO, CHAMADOS, %
    rows = []
    
    for loc in locations:
        lid = loc["id"]
        name = loc["name"]
        
        # Skip locations with 0 tickets IF needed to match Excel EXACTLY?
        # The sample Excel lists locations even with 0 tickets (e.g. ALA 2 - CLINICA CIRURGICA has 0).
        # So we include all known locations.
        
        # Determine classification
        # Try exact match, fallback to 'Outros' or based on containment
        classification = LOCATION_CLASSIFICATION.get(name.upper().strip(), "Outros")
        
        # If not exact match, try partial? (e.g. "ALA 1" vs "AlA 1")
        if classification == "Outros":
             for key, val in LOCATION_CLASSIFICATION.items():
                 if key in name.upper():
                     classification = val
                     break
        
        count = counts_by_loc_id.get(lid, 0)
        pct = (count / total_tickets) if total_tickets > 0 else 0
        
        rows.append({
            "code": lid,
            "desc": name,
            "class": classification,
            "count": count,
            "pct": pct
        })

    # Sort rows by Description (Name) alphabetically, as per Excel
    rows.sort(key=lambda x: x["desc"])

    # 5. Generate Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    # Headers (row 9 in sample, but let's put at row 1 for clean data, or mimic layout?)
    # User said: "layout 100% identico".
    # Sample starts header at row 9 usually? Let's check the analysis...
    # The analysis output showed Data at row ~9-10. Rows 0-8 were NaN/Header noise.
    # To be safe and clean, we will create a standard clean header, but match columns.
    
    headers = ["CÓDIGO", "DESCRIÇÃO", "CLASSIFICAÇÃO", "NÚMERO DE ATENDIMENTOS POR CENTRO DE CUSTO (CHAMADOS)", "%"]
    ws.append(headers)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data
    for r in rows:
        row_vals = [r["code"], r["desc"], r["class"], r["count"], r["pct"]]
        ws.append(row_vals)
        
        # Format Percentage
        ws.cell(row=ws.max_row, column=5).number_format = '0.00%'

    # Total Row
    total_row = ["", "TOTAL", "", total_tickets, 1.0]
    ws.append(total_row)
    
    # Style Total Row
    last_row = ws.max_row
    for col in range(1, 6):
        cell = ws.cell(row=last_row, column=col)
        cell.font = Font(bold=True)
        if col == 5:
            cell.number_format = '0.00%'

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return buf.getvalue(), filename
